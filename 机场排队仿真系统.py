import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import numpy as np
from collections import defaultdict, deque
import copy

class AirportQueueSimulator:
    def __init__(self, departure_time=20, arrival_time=10, num_runways=1):
        """
        初始化机场排队仿真器
        
        Args:
            departure_time: 出港时间（分钟），默认20分钟
            arrival_time: 入港时间（分钟），默认10分钟  
            num_runways: 跑道数量，默认1条
        """
        self.departure_time = departure_time  # 出港时间（分钟）
        self.arrival_time = arrival_time      # 入港时间（分钟）
        self.num_runways = num_runways        # 跑道数量
        
        # 机场跑道队列：{机场代码: {'departure': [], 'arrival': []}}
        self.airport_queues = defaultdict(lambda: {'departure': [], 'arrival': []})
        
        # 飞行计划数据
        self.flight_plans = {}
        
        # 仿真结果记录
        self.simulation_results = []
        
    def parse_time_string(self, time_str):
        """解析时间字符串，支持超过24小时的格式"""
        if isinstance(time_str, str):
            parts = time_str.split(':')
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2]) if len(parts) > 2 else 0
            
            # 转换为从0点开始的总分钟数
            total_minutes = hours * 60 + minutes + seconds / 60
            return total_minutes
        return 0
    
    def minutes_to_time_string(self, total_minutes):
        """将总分钟数转换回时间字符串格式"""
        hours = int(total_minutes // 60)
        minutes = int(total_minutes % 60)
        seconds = int((total_minutes % 1) * 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    def load_flight_plans(self, xml_file_path):
        """加载飞行计划XML文件"""
        print(f"正在加载飞行计划文件: {xml_file_path}")
        
        try:
            tree = ET.parse(xml_file_path)
            root = tree.getroot()
            
            for person in root.findall('person'):
                aircraft_id = person.get('id')
                plan = person.find('plan')
                
                if plan is not None:
                    activities = []
                    
                    for activity in plan.findall('act'):
                        act_type = activity.get('type')
                        link_id = activity.get('link')
                        start_time = activity.get('start_time')
                        end_time = activity.get('end_time')
                        x = float(activity.get('x', 0))
                        y = float(activity.get('y', 0))
                        
                        # 解析航站信息
                        if link_id and '-' in link_id:
                            airports = link_id.split('-')
                            origin = airports[0] if len(airports) > 0 else ''
                            destination = airports[1] if len(airports) > 1 else ''
                        else:
                            origin = ''
                            destination = ''
                        
                        activities.append({
                            'type': act_type,
                            'link': link_id,
                            'origin': origin,
                            'destination': destination,
                            'start_time': start_time,
                            'end_time': end_time,
                            'start_minutes': self.parse_time_string(start_time) if start_time else 0,
                            'end_minutes': self.parse_time_string(end_time) if end_time else 0,
                            'x': x,
                            'y': y
                        })
                    
                    self.flight_plans[aircraft_id] = activities
            
            print(f"成功加载 {len(self.flight_plans)} 架飞机的飞行计划")
            
        except Exception as e:
            print(f"加载飞行计划文件时出错: {e}")
    
    def collect_airport_activities(self):
        """收集每个机场的起降活动"""
        print("正在收集机场起降活动...")
        
        airport_activities = defaultdict(lambda: {'departure': [], 'arrival': []})
        
        for aircraft_id, activities in self.flight_plans.items():
            for i, activity in enumerate(activities):
                if activity['type'] == 'p':  # 出港准备
                    # p活动: link="A-B" 表示从A机场出发，所以p活动发生在A机场(origin)
                    airport = activity['origin']
                    if airport:
                        airport_activities[airport]['departure'].append({
                            'aircraft_id': aircraft_id,
                            'activity_index': i,
                            'scheduled_time': activity['start_minutes'],
                            'original_start': activity['start_minutes'],
                            'original_end': activity['end_minutes'],
                            'activity': activity
                        })
                
                elif activity['type'] == 'd':  # 入港降落
                    # d活动: link="A-B" 表示从A飞到B降落，所以d活动发生在B机场(destination)
                    airport = activity['destination']
                    if airport:
                        airport_activities[airport]['arrival'].append({
                            'aircraft_id': aircraft_id,
                            'activity_index': i,
                            'scheduled_time': activity['start_minutes'],
                            'original_start': activity['start_minutes'],
                            'original_end': activity['end_minutes'],
                            'activity': activity
                        })
        
        # 按计划时间排序
        for airport in airport_activities:
            airport_activities[airport]['departure'].sort(key=lambda x: x['scheduled_time'])
            airport_activities[airport]['arrival'].sort(key=lambda x: x['scheduled_time'])
        
        print(f"收集到 {len(airport_activities)} 个机场的活动")
        for airport, activities in airport_activities.items():
            print(f"  {airport}: {len(activities['departure'])}次出港, {len(activities['arrival'])}次入港")
        
        return airport_activities
    
    def simulate_queue(self, airport_activities, time_disturbance=0):
        """
        仿真机场排队系统（支持多跑道，出港入港分离）
        
        Args:
            airport_activities: 机场活动数据
            time_disturbance: 时间扰动（分钟），负数表示提前，正数表示延误
        """
        print("开始排队仿真...")
        
        # 初始化仿真结果
        self.simulation_results = []
        updated_plans = copy.deepcopy(self.flight_plans)
        
        # 对每个机场进行排队仿真
        for airport, activities in airport_activities.items():
            print(f"\n处理机场 {airport}:")
            
            # 出港队列仿真（独立的出港跑道）
            if activities['departure']:
                print(f"  处理 {len(activities['departure'])} 次出港 ({self.num_runways}条出港跑道)")
                departure_runway_end_times = [0] * self.num_runways  # 出港跑道结束时间
                
                for i, dep_activity in enumerate(activities['departure']):
                    aircraft_id = dep_activity['aircraft_id']
                    activity_index = dep_activity['activity_index']
                    scheduled_time = dep_activity['scheduled_time']
                    
                    # 选择最早可用的出港跑道
                    runway_idx = min(range(self.num_runways), key=lambda x: departure_runway_end_times[x])
                    earliest_available = departure_runway_end_times[runway_idx]
                    
                    # 计算实际开始时间（不能早于计划时间，也不能早于跑道可用时间）
                    actual_start = max(scheduled_time, earliest_available)
                    actual_end = actual_start + self.departure_time + time_disturbance
                    
                    # 更新出港跑道结束时间
                    departure_runway_end_times[runway_idx] = actual_end
                    
                    # 更新计划
                    updated_plans[aircraft_id][activity_index]['start_minutes'] = actual_start
                    updated_plans[aircraft_id][activity_index]['end_minutes'] = actual_end
                    updated_plans[aircraft_id][activity_index]['start_time'] = self.minutes_to_time_string(actual_start)
                    updated_plans[aircraft_id][activity_index]['end_time'] = self.minutes_to_time_string(actual_end)
                    
                    # 记录仿真结果（出港跑道编号从D1开始）
                    delay = actual_start - scheduled_time
                    self.simulation_results.append({
                        'aircraft_id': aircraft_id,
                        'airport': airport,
                        'activity_type': 'departure',
                        'scheduled_start': self.minutes_to_time_string(scheduled_time),
                        'actual_start': self.minutes_to_time_string(actual_start),
                        'actual_end': self.minutes_to_time_string(actual_end),
                        'delay_minutes': delay,
                        'queue_position': i + 1,
                        'runway': f'D{runway_idx + 1}'  # D1, D2, D3... 表示出港跑道
                    })
                    
                    if i < 8:  # 显示前8个的详细信息
                        print(f"    {aircraft_id}: 计划{self.minutes_to_time_string(scheduled_time)} -> "
                              f"实际{self.minutes_to_time_string(actual_start)}-{self.minutes_to_time_string(actual_end)} "
                              f"(延误{delay:.1f}分钟, 出港跑道D{runway_idx+1})")
            
            # 入港队列仿真（独立的入港跑道）
            if activities['arrival']:
                print(f"  处理 {len(activities['arrival'])} 次入港 ({self.num_runways}条入港跑道)")
                arrival_runway_end_times = [0] * self.num_runways  # 入港跑道结束时间
                
                for i, arr_activity in enumerate(activities['arrival']):
                    aircraft_id = arr_activity['aircraft_id']
                    activity_index = arr_activity['activity_index']
                    scheduled_time = arr_activity['scheduled_time']
                    
                    # 选择最早可用的入港跑道
                    runway_idx = min(range(self.num_runways), key=lambda x: arrival_runway_end_times[x])
                    earliest_available = arrival_runway_end_times[runway_idx]
                    
                    # 计算实际开始时间
                    actual_start = max(scheduled_time, earliest_available)
                    actual_end = actual_start + self.arrival_time + time_disturbance
                    
                    # 更新入港跑道结束时间
                    arrival_runway_end_times[runway_idx] = actual_end
                    
                    # 更新计划
                    updated_plans[aircraft_id][activity_index]['start_minutes'] = actual_start
                    updated_plans[aircraft_id][activity_index]['end_minutes'] = actual_end
                    updated_plans[aircraft_id][activity_index]['start_time'] = self.minutes_to_time_string(actual_start)
                    updated_plans[aircraft_id][activity_index]['end_time'] = self.minutes_to_time_string(actual_end)
                    
                    # 记录仿真结果（入港跑道编号从A1开始）
                    delay = actual_start - scheduled_time
                    self.simulation_results.append({
                        'aircraft_id': aircraft_id,
                        'airport': airport,
                        'activity_type': 'arrival',
                        'scheduled_start': self.minutes_to_time_string(scheduled_time),
                        'actual_start': self.minutes_to_time_string(actual_start),
                        'actual_end': self.minutes_to_time_string(actual_end),
                        'delay_minutes': delay,
                        'queue_position': i + 1,
                        'runway': f'A{runway_idx + 1}'  # A1, A2, A3... 表示入港跑道
                    })
                    
                    if i < 8:  # 显示前8个的详细信息
                        print(f"    {aircraft_id}: 计划{self.minutes_to_time_string(scheduled_time)} -> "
                              f"实际{self.minutes_to_time_string(actual_start)}-{self.minutes_to_time_string(actual_end)} "
                              f"(延误{delay:.1f}分钟, 入港跑道A{runway_idx+1})")
        
        return updated_plans
    
    def update_connected_activities(self, updated_plans):
        """更新相关联的活动时间（f、w等）"""
        print("正在更新相关活动时间...")
        
        for aircraft_id, activities in updated_plans.items():
            for i in range(len(activities)):
                activity = activities[i]
                
                # 如果当前活动是p（出港准备），更新后续的f（飞行）活动
                if activity['type'] == 'p' and i + 2 < len(activities):
                    flight_activity = activities[i + 2]  # p -> leg -> f
                    if flight_activity['type'] == 'f':
                        # f活动的结束时间 = p活动的结束时间
                        flight_activity['end_minutes'] = activity['end_minutes']
                        flight_activity['end_time'] = activity['end_time']
                
                # 如果当前活动是d（入港降落），更新后续的w（等待）活动
                if activity['type'] == 'd' and i + 2 < len(activities):
                    wait_activity = activities[i + 2]  # d -> leg -> w
                    if wait_activity['type'] == 'w':
                        # w活动的开始时间 = d活动的结束时间
                        wait_activity['start_minutes'] = activity['end_minutes']
                        wait_activity['start_time'] = activity['end_time']
                
                # 更新leg（移动）的时间
                if activity['type'] == 'leg':
                    # leg通常在两个活动之间，时间为0
                    if i > 0:
                        prev_activity = activities[i - 1]
                        activity['dep_time'] = prev_activity['end_time']
        
        return updated_plans
    
    def save_updated_plans(self, updated_plans, output_file):
        """保存更新后的飞行计划到XML文件"""
        print(f"正在保存更新后的飞行计划到: {output_file}")
        
        # 创建XML根元素
        root = ET.Element('plans')
        root.set('xml:lang', 'de-CH')
        
        # 添加DTD声明
        root_str = '<?xml version="1.0" ?>\n<!DOCTYPE plans SYSTEM "http://www.matsim.org/files/dtd/plans_v4.dtd">\n'
        
        for aircraft_id, activities in updated_plans.items():
            person_elem = ET.SubElement(root, 'person')
            person_elem.set('id', aircraft_id)
            
            plan_elem = ET.SubElement(person_elem, 'plan')
            
            for i, activity in enumerate(activities):
                if activity['type'] in ['p', 'f', 'd', 'w']:
                    # 创建活动元素
                    act_elem = ET.SubElement(plan_elem, 'act')
                    act_elem.set('type', activity['type'])
                    act_elem.set('x', str(activity['x']))
                    act_elem.set('y', str(activity['y']))
                    act_elem.set('link', activity['link'])
                    
                    if activity['start_time']:
                        act_elem.set('start_time', activity['start_time'])
                    if activity['end_time']:
                        act_elem.set('end_time', activity['end_time'])
                
                # 在活动之间添加leg元素
                if i < len(activities) - 1:
                    leg_elem = ET.SubElement(plan_elem, 'leg')
                    leg_elem.set('mode', 'car')
                    
                    next_activity = activities[i + 1]
                    if activity['type'] == 'f':
                        # 飞行leg有出发时间
                        leg_elem.set('dep_time', activity['end_time'])
                    else:
                        # 其他leg时间为0
                        leg_elem.set('trav_time', '00:00:00')
        
        # 格式化并保存XML
        tree = ET.ElementTree(root)
        ET.indent(tree, space="\t", level=0)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(root_str)
            tree.write(f, encoding='unicode', xml_declaration=False)
        
        print(f"✅ 更新后的飞行计划已保存")
    
    def generate_analysis_report(self, original_data_file, output_excel):
        """生成分析报告Excel文件 - 基于原始航班数据添加排队仿真结果"""
        print(f"正在生成分析报告...")
        
        try:
            # 读取原始航班数据
            df_original = pd.read_excel(original_data_file)
            print(f"读取原始数据: {len(df_original)} 条记录")
            
            # 转换仿真结果为DataFrame
            df_simulation = pd.DataFrame(self.simulation_results)
            print(f"仿真结果: {len(df_simulation)} 条记录")
            
            # 复制原始数据作为基础
            df_enhanced = df_original.copy()
            
            # 初始化新列
            new_columns = [
                '出港排队延误_分钟', '出港排队位置', '出港使用跑道',
                '入港排队延误_分钟', '入港排队位置', '入港使用跑道',
                '仿真出港开始时间', '仿真出港结束时间',
                '仿真入港开始时间', '仿真入港结束时间',
                '是否有排队数据'
            ]
            
            for col in new_columns:
                df_enhanced[col] = None
            
            # 为每条原始航班数据匹配仿真结果
            matched_count = 0
            print("正在匹配仿真数据...")
            
            # 使用批处理来避免超时
            batch_size = 10000
            total_batches = (len(df_original) + batch_size - 1) // batch_size
            
            for batch_idx in range(total_batches):
                start_idx = batch_idx * batch_size
                end_idx = min((batch_idx + 1) * batch_size, len(df_original))
                
                print(f"  处理批次 {batch_idx + 1}/{total_batches}: 行 {start_idx}-{end_idx}")
                
                for idx in range(start_idx, end_idx):
                    row = df_original.iloc[idx]
                    
                    # 提取航班信息
                    tail_number = str(row['机尾号'])  # 机尾号对应aircraft_id
                    departure_airport = str(row['计划起飞站四字码'])
                    arrival_airport = str(row['计划到达站四字码'])
                    
                    # 查找该飞机的仿真数据
                    aircraft_sim = df_simulation[df_simulation['aircraft_id'] == tail_number]
                    
                    if len(aircraft_sim) == 0:
                        df_enhanced.loc[idx, '是否有排队数据'] = '无仿真数据'
                        continue
                    
                    # 查找出港数据
                    departure_sim = aircraft_sim[
                        (aircraft_sim['activity_type'] == 'departure') & 
                        (aircraft_sim['airport'] == departure_airport)
                    ]
                    
                    # 查找入港数据
                    arrival_sim = aircraft_sim[
                        (aircraft_sim['activity_type'] == 'arrival') & 
                        (aircraft_sim['airport'] == arrival_airport)
                    ]
                    
                    # 如果找到了出港数据
                    if len(departure_sim) > 0:
                        dep_data = departure_sim.iloc[0]  # 取第一条匹配的记录
                        df_enhanced.loc[idx, '出港排队延误_分钟'] = dep_data['delay_minutes']
                        df_enhanced.loc[idx, '出港排队位置'] = dep_data['queue_position']
                        df_enhanced.loc[idx, '出港使用跑道'] = dep_data['runway']
                        df_enhanced.loc[idx, '仿真出港开始时间'] = dep_data['actual_start']
                        df_enhanced.loc[idx, '仿真出港结束时间'] = dep_data['actual_end']
                    
                    # 如果找到了入港数据
                    if len(arrival_sim) > 0:
                        arr_data = arrival_sim.iloc[0]  # 取第一条匹配的记录
                        df_enhanced.loc[idx, '入港排队延误_分钟'] = arr_data['delay_minutes']
                        df_enhanced.loc[idx, '入港排队位置'] = arr_data['queue_position']
                        df_enhanced.loc[idx, '入港使用跑道'] = arr_data['runway']
                        df_enhanced.loc[idx, '仿真入港开始时间'] = arr_data['actual_start']
                        df_enhanced.loc[idx, '仿真入港结束时间'] = arr_data['actual_end']
                    
                    # 标记数据状态
                    if len(departure_sim) > 0 and len(arrival_sim) > 0:
                        df_enhanced.loc[idx, '是否有排队数据'] = '完整数据'
                        matched_count += 1
                    elif len(departure_sim) > 0 or len(arrival_sim) > 0:
                        df_enhanced.loc[idx, '是否有排队数据'] = '部分数据'
                        matched_count += 1
                    else:
                        df_enhanced.loc[idx, '是否有排队数据'] = '未匹配'
            
            print("数据匹配完成，正在生成统计...")
            
            # 生成统计分析
            stats_data = []
            
            # 基本统计
            total_flights = len(df_enhanced)
            complete_data = len(df_enhanced[df_enhanced['是否有排队数据'] == '完整数据'])
            partial_data = len(df_enhanced[df_enhanced['是否有排队数据'] == '部分数据'])
            no_data = len(df_enhanced[df_enhanced['是否有排队数据'] == '无仿真数据'])
            
            stats_data.append(['总航班数', total_flights])
            stats_data.append(['有完整排队数据', complete_data])
            stats_data.append(['有部分排队数据', partial_data])
            stats_data.append(['无排队数据', no_data])
            stats_data.append(['数据匹配率(%)', f"{matched_count/total_flights*100:.1f}"])
            
            # 延误统计
            dep_delays = df_enhanced['出港排队延误_分钟'].dropna()
            arr_delays = df_enhanced['入港排队延误_分钟'].dropna()
            
            if len(dep_delays) > 0:
                stats_data.append(['平均出港排队延误(分钟)', f"{dep_delays.mean():.2f}"])
                stats_data.append(['最大出港排队延误(分钟)', f"{dep_delays.max():.2f}"])
                stats_data.append(['出港延误航班比例(%)', f"{(dep_delays > 0).mean()*100:.1f}"])
            
            if len(arr_delays) > 0:
                stats_data.append(['平均入港排队延误(分钟)', f"{arr_delays.mean():.2f}"])
                stats_data.append(['最大入港排队延误(分钟)', f"{arr_delays.max():.2f}"])
                stats_data.append(['入港延误航班比例(%)', f"{(arr_delays > 0).mean()*100:.1f}"])
            
            df_stats = pd.DataFrame(stats_data, columns=['指标', '数值'])
            
            print("正在保存Excel文件...")
            
            # 分块保存到Excel - 避免超时
            try:
                # 先保存统计分析和仿真结果（较小的数据）
                print("  保存统计分析和仿真结果...")
                with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                    df_stats.to_excel(writer, sheet_name='统计分析', index=False)
                    if len(df_simulation) > 0:
                        df_simulation.to_excel(writer, sheet_name='原始仿真结果', index=False)
                
                # 保存主要数据 - 使用更高效的方法
                print("  保存增强航班数据...")
                
                # 先保存到CSV（更快），然后转换为Excel
                csv_temp_file = output_excel.replace('.xlsx', '_temp.csv')
                df_enhanced.to_csv(csv_temp_file, index=False, encoding='utf-8-sig')
                
                # 重新读取CSV并添加到Excel
                df_temp = pd.read_csv(csv_temp_file, encoding='utf-8-sig')
                
                # 使用xlsxwriter引擎（更快）
                try:
                    with pd.ExcelWriter(output_excel, engine='xlsxwriter', mode='a') as writer:
                        df_temp.to_excel(writer, sheet_name='增强航班数据', index=False)
                except:
                    # 如果xlsxwriter失败，使用openpyxl但分批写入
                    print("  使用分批写入模式...")
                    import openpyxl
                    
                    # 加载现有工作簿
                    wb = openpyxl.load_workbook(output_excel)
                    ws = wb.create_sheet('增强航班数据')
                    
                    # 写入列标题
                    for col_idx, col_name in enumerate(df_enhanced.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    
                    # 分批写入数据
                    batch_size = 1000
                    for batch_start in range(0, len(df_enhanced), batch_size):
                        batch_end = min(batch_start + batch_size, len(df_enhanced))
                        print(f"    写入行 {batch_start+1}-{batch_end}")
                        
                        for row_idx, (_, row) in enumerate(df_enhanced.iloc[batch_start:batch_end].iterrows(), batch_start + 2):
                            for col_idx, value in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    wb.save(output_excel)
                    wb.close()
                
                # 清理临时文件
                import os
                if os.path.exists(csv_temp_file):
                    os.remove(csv_temp_file)
                
                print(f"✅ 增强分析报告已保存到: {output_excel}")
                
            except Exception as save_error:
                print(f"保存Excel时出错: {save_error}")
                # 备用方案：保存为CSV文件
                csv_output = output_excel.replace('.xlsx', '_backup.csv')
                print(f"备用方案：保存为CSV文件: {csv_output}")
                df_enhanced.to_csv(csv_output, index=False, encoding='utf-8-sig')
                
                # 保存统计数据
                stats_output = output_excel.replace('.xlsx', '_stats.csv')
                df_stats.to_csv(stats_output, index=False, encoding='utf-8-sig')
                print(f"统计数据已保存到: {stats_output}")
            
            print(f"   - 总航班记录: {total_flights} 条")
            print(f"   - 匹配到排队数据: {matched_count} 条 ({matched_count/total_flights*100:.1f}%)")
            print(f"   - 完整数据: {complete_data} 条")
            print(f"   - 部分数据: {partial_data} 条")
            if len(dep_delays) > 0:
                print(f"   - 平均出港排队延误: {dep_delays.mean():.1f} 分钟")
            if len(arr_delays) > 0:
                print(f"   - 平均入港排队延误: {arr_delays.mean():.1f} 分钟")
            
        except Exception as e:
            print(f"生成分析报告时出错: {e}")
            import traceback
            traceback.print_exc()


def main():
    """主函数"""
    print("=== 机场排队仿真系统 ===")
    
    # 文件路径
    flight_plans_file = "/Users/megrez/Library/Mobile Documents/com~apple~CloudDocs/BUAA/科研/挑战杯/航空挑战杯/仿真/all_flight_plans.xml"
    original_data_file = "/Users/megrez/Library/Mobile Documents/com~apple~CloudDocs/BUAA/科研/挑战杯/航空挑战杯/数据/三段式飞行- plan所需数据.xlsx"
    
    output_plans_file = "/Users/megrez/Library/Mobile Documents/com~apple~CloudDocs/BUAA/科研/挑战杯/航空挑战杯/仿真/updated_all_flight_plans.xml"
    output_report_file = "/Users/megrez/Library/Mobile Documents/com~apple~CloudDocs/BUAA/科研/挑战杯/航空挑战杯/完整机场排队仿真分析报告.xlsx"
    
    # 创建仿真器
    simulator = AirportQueueSimulator(
        departure_time=20,  # 出港时间20分钟
        arrival_time=10,    # 入港时间10分钟
        num_runways=10       # 每个机场1条跑道
    )
    
    # 加载飞行计划
    simulator.load_flight_plans(flight_plans_file)
    
    # 收集机场活动
    airport_activities = simulator.collect_airport_activities()
    
    # 执行排队仿真
    updated_plans = simulator.simulate_queue(airport_activities, time_disturbance=0)
    
    # 更新相关活动
    updated_plans = simulator.update_connected_activities(updated_plans)
    
    # 保存更新后的计划
    simulator.save_updated_plans(updated_plans, output_plans_file)
    
    # 生成分析报告
    simulator.generate_analysis_report(original_data_file, output_report_file)
    
    print("\n=== 仿真完成 ===")


if __name__ == "__main__":
    main()
